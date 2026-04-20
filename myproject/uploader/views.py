# uploader/views.py
import pandas as pd
from django.db import connection, models as django_models
from django.apps import apps
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .forms import UploadForm
from .models import UploadedFile
from .services import (
    get_dynamic_model,
    build_month_filter,
    save_filtered_plan_table,
    get_materials_by_order, get_order_status
)


def home(request):
    return render(request, 'uploader/home.html')


def upload(request, file_type):
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            uploaded_file = UploadedFile.objects.create(type=file_type, file=file)
            process_excel_to_db(uploaded_file, file_type)
            return redirect('success', file_type=file_type)
    else:
        form = UploadForm()
    return render(request, 'uploader/upload.html', {'form': form, 'file_type': file_type})


def process_excel_to_db(uploaded_file, file_type):
    """Создаёт таблицу при первой загрузке + обновляет при изменении заголовков"""
    file_path = uploaded_file.file.path
    df = pd.read_excel(file_path)
    headers = df.columns.tolist()
    uploaded_file.headers = headers
    uploaded_file.save()

    app_label = 'uploader'
    model_name = file_type.capitalize()
    table_name = f"uploader_{file_type}"

    # Пытаемся получить модель
    try:
        DynamicModel = get_dynamic_model(file_type)
        existing_fields = [f.name for f in DynamicModel._meta.get_fields() if f.name != 'id']
        if sorted(existing_fields) != sorted(headers):
            raise LookupError("Структура изменилась")
        DynamicModel.objects.all().delete()
    except (ValueError, LookupError):
        # Таблицы нет или структура изменилась → создаём новую
        with connection.cursor() as cursor:
            columns_def = ', '.join([f'"{col}" TEXT' for col in headers])
            cursor.execute(f"""
                DROP TABLE IF EXISTS "{table_name}" CASCADE;
                CREATE TABLE "{table_name}" (
                    id SERIAL PRIMARY KEY,
                    {columns_def}
                );
            """)

        # Создаём модель
        fields = {'__module__': f'{app_label}.models'}
        for col in headers:
            fields[col] = django_models.CharField(max_length=500, blank=True, null=True)
        DynamicModel = type(model_name, (django_models.Model,), fields)
        DynamicModel._meta.db_table = table_name
        apps.register_model(app_label, DynamicModel)

    # Заполняем данными с корректной обработкой чисел
    for _, row in df.iterrows():
        data = {}
        for col in headers:
            val = row[col]
            if pd.isna(val):
                data[col] = None
            elif col == 'Количество':  # ← обрабатываем только это поле
                try:
                    # Убираем пробелы, запятые → превращаем в float
                    cleaned = str(val).replace(' ', '').replace(',', '.')
                    data[col] = float(cleaned)
                except (ValueError, TypeError):
                    data[col] = None  # если не число — сохраняем None
            elif isinstance(val, (int, float)):
                data[col] = float(val)  # все числа как float
            else:
                data[col] = str(val)

        DynamicModel.objects.create(**data)


def success(request, file_type):
    last_file = UploadedFile.objects.filter(type=file_type).order_by('-uploaded_at').first()
    return render(request, 'uploader/success.html', {
        'message': f'Файл для {file_type} успешно загружен!',
        'file_type': file_type,
        'uploaded_file': last_file
    })


def view_table(request, file_type):
    DynamicModel = get_dynamic_model(file_type)
    source_table_name = f"uploader_{file_type}"

    all_available_columns = sorted([
        f.name for f in DynamicModel._meta.get_fields()
        if f.name != 'id' and not f.name.startswith('_')
    ])

    is_plan = file_type.lower() == 'plan'

    if is_plan:
        default_cols = ['Заказ', 'ЗапланНачало', 'ПроизвУчасток']
        columns_for_view = [col for col in all_available_columns if col not in default_cols]

        selected_month_str = request.GET.get('month', '').strip().lower()
        selected_section = request.GET.get('section', '').strip()
        selected_columns_input = request.GET.getlist('columns')

        selected_columns = default_cols + [c for c in selected_columns_input if c in columns_for_view]
        if not selected_columns:
            selected_columns = default_cols if default_cols else all_available_columns

        # Фильтрация
        queryset = DynamicModel.objects.all()
        if selected_section and 'ПроизвУчасток' in all_available_columns:
            queryset = queryset.filter(ПроизвУчасток=selected_section)

        date_fields = ['БазисСрокНачала', 'БазисСрокКонца', 'ЗапланНачало']
        month_condition, month_params = build_month_filter(selected_month_str, date_fields)

        if month_condition or selected_section:
            with connection.cursor() as cursor:
                query = f'SELECT * FROM "{source_table_name}" WHERE 1=1'
                params = []
                if selected_section:
                    query += ' AND "ПроизвУчасток" = %s'
                    params.append(selected_section)
                if month_condition:
                    query += ' '+month_condition
                    params.extend(month_params)
                cursor.execute(query, params)
                rows = [DynamicModel(**dict(zip([f.name for f in DynamicModel._meta.get_fields()], row)))
                        for row in cursor.fetchall()]
        else:
            rows = list(queryset)

        saved_message = None
        if (selected_month_str or selected_section) and selected_columns:
            save_filtered_plan_table(selected_columns, selected_section, month_condition, month_params,
                                     source_table_name)
            saved_message = f'Таблица uploader_filtered_plan обновлена ({len(rows)} строк)'
    else:
        # Для всех остальных типов — все колонки, без фильтров
        selected_columns = all_available_columns
        selected_month_str = ''
        selected_section = ''
        rows = list(DynamicModel.objects.all())
        saved_message = None
        columns_for_view = []

    table_data = []
    for row in rows:
        row_dict = {col: getattr(row, col, None) for col in selected_columns}

        # Добавляем статус заказа
        status = 'no_materials'
        if 'Заказ' in row_dict and row_dict['Заказ']:
            status = get_order_status(str(row_dict['Заказ']).strip())

        row_dict['status'] = status  # ← обычное поле, без подчёркивания
        table_data.append(row_dict)

    sections = []
    if is_plan and 'ПроизвУчасток' in all_available_columns:
        sections = list(DynamicModel.objects
                        .values_list('ПроизвУчасток', flat=True)
                        .distinct()
                        .exclude(ПроизвУчасток__isnull=True)
                        .order_by('ПроизвУчасток'))

    months_list = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']

    context = {
        'file_type': file_type,
        'columns': selected_columns,
        'columns_for_view': columns_for_view if is_plan else [],
        'selected_columns': selected_columns,
        'table_data': table_data,
        'months': months_list if is_plan else None,
        'selected_month': selected_month_str.capitalize() if selected_month_str else None,
        'selected_month_lower': selected_month_str if is_plan else '',
        'sections': sections,
        'selected_section': selected_section,
        'row_count': len(table_data),
        'saved_message': saved_message,
    }

    return render(request, 'uploader/view_file.html', context)


def get_materials_for_order(request, order_number):
    """AJAX-эндпоинт: материалы по номеру заказа"""
    materials = get_materials_by_order(order_number)

    columns = list(materials[0].keys()) if materials else []

    return JsonResponse({
        'order_number': order_number,
        'columns': columns,
        'materials': materials,
        'count': len(materials)
    })

from openpyxl import Workbook
from datetime import datetime
from django.http import HttpResponse

def download_table_excel(request, file_type):
    """Скачивание таблицы в Excel. Приоритет — отфильтрованная таблица для plan"""
    try:
        if file_type.lower() == 'plan':
            table_name = 'uploader_filtered_plan'
        else:
            table_name = f"uploader_{file_type}"

        # Проверяем, существует ли таблица
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_schema = 'public' AND table_name = %s
                );
            """, [table_name])
            exists = cursor.fetchone()[0]

        # Если filtered_plan не существует — берём оригинальную таблицу plan
        if not exists and file_type.lower() == 'plan':
            table_name = 'uploader_plan'

        # Получаем данные
        if table_name == 'uploader_filtered_plan':
            # Для filtered_plan создаём временную модель
            with connection.cursor() as cursor:
                description = connection.introspection.get_table_description(cursor, table_name)
                columns = [field.name for field in description if field.name != 'id']

            fields = {'__module__': 'uploader.models'}
            for col in columns:
                fields[col] = django_models.CharField(max_length=500, blank=True, null=True)

            DynamicModel = type('FilteredPlan', (django_models.Model,), fields)
            DynamicModel._meta.db_table = table_name
            apps.register_model('uploader', DynamicModel)
        else:
            DynamicModel = get_dynamic_model(file_type)

        rows = DynamicModel.objects.all()

        # Создаём Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = file_type.upper()

        # Заголовки
        columns = [f.name for f in DynamicModel._meta.get_fields() if f.name != 'id']
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        # Заполняем данные
        for row_num, obj in enumerate(rows, 2):
            for col_num, col in enumerate(columns, 1):
                value = getattr(obj, col, '')
                ws.cell(row=row_num, column=col_num, value=value)

        # Ответ для скачивания
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{file_type}_table_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    except Exception as e:
        print(f"Ошибка скачивания Excel: {e}")
        return HttpResponse(f"Ошибка при формировании файла: {str(e)}", status=500)